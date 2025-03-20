import streamlit as st
import subprocess
import sys
import pandas as pd
import os
import base64
import re
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# 检查是否已安装rdkit
try:
    from rdkit import Chem
    from rdkit.Chem import Draw
except ImportError:
    st.error("""
    RDKit 未安装。请按照以下步骤安装：
    
    1. 在终端中运行：
    ```bash
    conda install -c conda-forge rdkit
    ```
    
    2. 安装完成后重新运行应用。
    """)
    st.stop()

from PIL import Image
import io
import zipfile

def is_valid_smiles(smiles):
    """检查字符串是否为有效的SMILES"""
    try:
        mol = Chem.MolFromSmiles(str(smiles))
        return mol is not None
    except:
        return False

def find_smiles_column(df):
    """自动识别包含SMILES的列"""
    for column in df.columns:
        # 检查列名是否包含smiles（不区分大小写）
        if 'smiles' in column.lower():
            # 验证该列是否包含有效的SMILES
            valid_count = sum(1 for x in df[column] if is_valid_smiles(x))
            if valid_count > len(df) * 0.5:  # 如果超过50%的值是有效的SMILES
                return column
    return None

def save_to_excel(df, smiles_column, output_path):
    """将数据保存到Excel文件，包含结构图片"""
    wb = Workbook()
    ws = wb.active
    
    # 写入列标题
    for col, header in enumerate(df.columns, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 创建临时目录存储图片
    if not os.path.exists('temp_images'):
        os.makedirs('temp_images')
    
    # 处理每一行数据
    for row_idx, row in df.iterrows():
        excel_row = row_idx + 2  # Excel行号从1开始，标题占第1行
        
        # 写入数据
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=excel_row, column=col_idx, value=value)
        
        # 处理SMILES并添加图片
        smiles = str(row[smiles_column])
        try:
            mol = Chem.MolFromSmiles(smiles)
            if mol is not None:
                # 生成分子图片
                img = Draw.MolToImage(mol, size=(400, 400))
                img_path = f'temp_images/molecule_{row_idx}.png'
                img.save(img_path)
                
                # 将图片添加到Excel
                img = XLImage(img_path)
                # 调整图片大小
                img.width = 200
                img.height = 200
                # 在SMILES列后面添加图片
                ws.add_image(img, f'{get_column_letter(len(df.columns) + 1)}{excel_row}')
        except Exception as e:
            st.warning(f"处理第 {row_idx + 1} 行的SMILES时出错: {str(e)}")
    
    # 调整列宽
    for col in range(1, len(df.columns) + 2):  # +2是因为多了一列图片
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # 保存Excel文件
    wb.save(output_path)
    
    # 清理临时文件
    for file in os.listdir('temp_images'):
        os.remove(os.path.join('temp_images', file))
    os.rmdir('temp_images')

st.set_page_config(page_title="SMILES 结构查看器", layout="wide")

st.title("SMILES 结构查看器")
st.write("输入 SMILES 字符串或上传 CSV 文件，查看对应的化学结构")

# 创建两个标签页
tab1, tab2 = st.tabs(["单个转换", "批量转换"])

with tab1:
    # 创建两列布局
    col1, col2 = st.columns(2)

    with col1:
        smiles_input = st.text_input("输入 SMILES 字符串", "C1=CC=CC=C1")
        
        if smiles_input:
            try:
                # 创建 RDKit 分子对象
                mol = Chem.MolFromSmiles(smiles_input)
                if mol is None:
                    st.error("无效的 SMILES 字符串")
                else:
                    # 生成分子图片
                    img = Draw.MolToImage(mol, size=(400, 400))
                    
                    # 将图片转换为字节流
                    img_byte_arr = io.BytesIO()
                    img.save(img_byte_arr, format='PNG')
                    img_byte_arr = img_byte_arr.getvalue()
                    
                    # 显示图片
                    st.image(img_byte_arr, caption="化学结构", use_container_width=True)
                    
                    # 添加下载按钮
                    st.download_button(
                        label="下载结构图片",
                        data=img_byte_arr,
                        file_name="molecule.png",
                        mime="image/png"
                    )
            except Exception as e:
                st.error(f"处理时出错: {str(e)}")

    with col2:
        st.markdown("""
        ### 使用说明
        1. 在左侧输入框中输入 SMILES 字符串
        2. 系统会自动显示对应的化学结构
        3. 可以点击"下载结构图片"保存结构图
        
        ### 示例 SMILES
        - 苯: `C1=CC=CC=C1`
        - 乙醇: `CCO`
        - 阿司匹林: `CC(=O)OC1=CC=CC=C1C(=O)O`
        """)

with tab2:
    st.markdown("""
    ### 批量转换说明
    1. 上传包含 SMILES 列的 CSV 文件
    2. 系统会自动识别包含 SMILES 的列
    3. 系统会自动生成并显示所有结构的图片
    4. 可以下载单个结构图片或导出到Excel
    """)
    
    uploaded_file = st.file_uploader("上传 CSV 文件", type=['csv'])
    
    if uploaded_file is not None:
        try:
            # 读取CSV文件
            df = pd.read_csv(uploaded_file)
            
            # 显示数据预览
            st.subheader("数据预览")
            st.dataframe(df.head())
            
            # 自动识别SMILES列
            smiles_column = find_smiles_column(df)
            
            if smiles_column:
                st.success(f"自动识别到 SMILES 列: {smiles_column}")
            else:
                smiles_column = st.selectbox("未自动识别到 SMILES 列，请手动选择", df.columns)
            
            if smiles_column:
                # 创建两列布局用于显示图片
                col1, col2 = st.columns(2)
                
                # 处理每个SMILES
                success_count = 0
                error_count = 0
                error_smiles = []
                
                for idx, row in df.iterrows():
                    smiles = str(row[smiles_column])
                    try:
                        mol = Chem.MolFromSmiles(smiles)
                        if mol is not None:
                            # 生成分子图片
                            img = Draw.MolToImage(mol, size=(400, 400))
                            
                            # 将图片转换为字节流
                            img_byte_arr = io.BytesIO()
                            img.save(img_byte_arr, format='PNG')
                            img_byte_arr = img_byte_arr.getvalue()
                            
                            # 创建其他列的数据显示
                            other_data = {col: row[col] for col in df.columns if col != smiles_column}
                            data_text = "\n".join([f"{k}: {v}" for k, v in other_data.items()])
                            
                            # 在对应的列中显示图片
                            if idx % 2 == 0:
                                with col1:
                                    st.image(img_byte_arr, caption=f"SMILES: {smiles}\n{data_text}", use_container_width=True)
                                    st.download_button(
                                        label=f"下载结构 {idx+1}",
                                        data=img_byte_arr,
                                        file_name=f"molecule_{idx+1}.png",
                                        mime="image/png",
                                        key=f"download_{idx}"
                                    )
                            else:
                                with col2:
                                    st.image(img_byte_arr, caption=f"SMILES: {smiles}\n{data_text}", use_container_width=True)
                                    st.download_button(
                                        label=f"下载结构 {idx+1}",
                                        data=img_byte_arr,
                                        file_name=f"molecule_{idx+1}.png",
                                        mime="image/png",
                                        key=f"download_{idx}"
                                    )
                            
                            success_count += 1
                        else:
                            error_count += 1
                            error_smiles.append(smiles)
                    except Exception as e:
                        error_count += 1
                        error_smiles.append(smiles)
                
                # 显示统计信息
                st.success(f"处理完成！成功: {success_count}, 失败: {error_count}")
                
                # 如果有失败的SMILES，显示错误信息
                if error_smiles:
                    with st.expander("查看失败的 SMILES"):
                        st.write("以下 SMILES 处理失败：")
                        for smiles in error_smiles:
                            st.code(smiles)
                
                # 添加导出到Excel的功能
                if st.button("导出到Excel"):
                    try:
                        # 创建临时Excel文件
                        excel_path = "molecules_with_structures.xlsx"
                        save_to_excel(df, smiles_column, excel_path)
                        
                        # 读取Excel文件并创建下载按钮
                        with open(excel_path, 'rb') as f:
                            st.download_button(
                                label="下载Excel文件",
                                data=f,
                                file_name="molecules_with_structures.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        
                        # 删除临时文件
                        os.remove(excel_path)
                    except Exception as e:
                        st.error(f"导出Excel时出错: {str(e)}")
                
        except Exception as e:
            st.error(f"处理CSV文件时出错: {str(e)}") 